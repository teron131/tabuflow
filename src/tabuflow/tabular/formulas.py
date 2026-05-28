"""Formula metadata helpers for workbook-backed tabular extraction."""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path
import re
from typing import Any

from openpyxl.formula import Tokenizer
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter


def _cell_text(value: Any) -> str:
    """Return the same text shape used by tabular row loading."""
    return "" if value is None else str(value)


CELL_REFERENCE_RE = re.compile(r"^(?P<absolute_column>\$?)(?P<column>[A-Za-z]{1,3})(?P<absolute_row>\$?)(?P<row>[1-9][0-9]*)$")
DYNAMIC_REFERENCE_FUNCTIONS = {"INDIRECT", "OFFSET"}


def _unquote_scope(scope: str) -> str:
    """Return an Excel sheet/workbook scope without surrounding quotes."""
    if len(scope) >= 2 and scope.startswith("'") and scope.endswith("'"):
        return scope[1:-1].replace("''", "'")
    return scope


def _split_reference_scope(raw_reference: str) -> tuple[str | None, str]:
    """Split a tokenized Excel reference into scope and cell/range text."""
    if "!" not in raw_reference:
        return None, raw_reference
    scope, address = raw_reference.rsplit("!", 1)
    return scope, address


def _scope_parts(raw_scope: str | None) -> tuple[str | None, str | None]:
    """Return workbook and sheet names from an optional Excel reference scope."""
    if raw_scope is None:
        return None, None

    scope = _unquote_scope(raw_scope)
    if scope.startswith("["):
        workbook_end = scope.find("]")
        if workbook_end >= 0:
            workbook = scope[: workbook_end + 1]
            sheet = scope[workbook_end + 1 :] or None
            return workbook, sheet

    return None, scope or None


def _reference_endpoint(endpoint: str) -> dict[str, Any] | None:
    """Return structured A1 coordinates for one cell endpoint."""
    match = CELL_REFERENCE_RE.fullmatch(endpoint)
    if match is None:
        return None

    row = int(match.group("row"))
    column = column_index_from_string(match.group("column"))
    return {
        "row": row,
        "column": column,
        "coordinate": f"{get_column_letter(column)}{row}",
        "absolute_row": bool(match.group("absolute_row")),
        "absolute_column": bool(match.group("absolute_column")),
    }


def _direct_reference(raw_reference: str) -> dict[str, Any] | None:
    """Parse a direct A1 cell or range reference, if the token is one."""
    raw_scope, address = _split_reference_scope(raw_reference)
    parts = address.split(":")
    if len(parts) > 2:
        return None

    start = _reference_endpoint(parts[0])
    end = _reference_endpoint(parts[-1])
    if start is None or end is None:
        return None

    workbook, sheet = _scope_parts(raw_scope)
    return {
        "kind": "cell" if start["coordinate"] == end["coordinate"] else "range",
        "raw": raw_reference,
        "workbook": workbook,
        "sheet": sheet,
        "start": start,
        "end": end,
    }


def parse_formula_references(formula: str) -> dict[str, list[dict[str, Any]]]:
    """Return direct formula references plus unresolved static-analysis hints."""
    references: list[dict[str, Any]] = []
    unresolved_references: list[dict[str, Any]] = []

    try:
        tokens = Tokenizer(formula).items
    except Exception as exc:
        return {
            "references": [],
            "unresolved_references": [{"kind": "parse_error", "raw": formula, "message": str(exc)}],
        }

    for token in tokens:
        if token.type == "OPERAND" and token.subtype == "RANGE":
            reference = _direct_reference(token.value)
            if reference is None:
                unresolved_references.append({"kind": "range_token", "raw": token.value})
            else:
                references.append(reference)
            continue

        if token.type == "FUNC" and token.subtype == "OPEN":
            function_name = token.value[:-1].upper()
            if function_name in DYNAMIC_REFERENCE_FUNCTIONS:
                unresolved_references.append({"kind": "dynamic_function", "raw": function_name})

    return {
        "references": references,
        "unresolved_references": unresolved_references,
    }


def workbook_formula_cells(
    path: str | Path,
    *,
    sheet: str | None = None,
) -> list[dict[str, Any]]:
    """Return formula cells pinned to absolute worksheet coordinates."""
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        return []

    formula_workbook = load_workbook(path, read_only=False, data_only=False)
    value_workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet_name = sheet or formula_workbook.sheetnames[0]
        if sheet_name not in formula_workbook.sheetnames:
            raise ValueError(f"Unknown worksheet '{sheet_name}' in {path.name}")
        formula_sheet = formula_workbook[sheet_name]
        value_sheet = value_workbook[sheet_name]
        formulas: list[dict[str, Any]] = []
        for row in formula_sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                parsed_references = parse_formula_references(cell.value)
                formulas.append(
                    {
                        "row": cell.row,
                        "column": cell.column,
                        "coordinate": cell.coordinate,
                        "formula": cell.value,
                        "cached_value": _cell_text(value_sheet.cell(cell.row, cell.column).value),
                        **parsed_references,
                    }
                )
        return formulas
    finally:
        formula_workbook.close()
        value_workbook.close()


def formulas_in_window(
    formulas: list[dict[str, Any]],
    *,
    row_start: int,
    row_end: int,
    column_start: int,
    column_end: int,
) -> list[dict[str, Any]]:
    """Return formula cells that intersect a bounded sheet window."""
    return [formula for formula in formulas if row_start <= int(formula["row"]) <= row_end and column_start <= int(formula["column"]) <= column_end]


def formula_table_refs(
    formula: dict[str, Any],
    *,
    recovered_tables: list[dict[str, Any]],
    loaded_tables: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Return table-relative overlays for one formula cell."""
    row = int(formula["row"])
    column = int(formula["column"])
    refs: list[dict[str, Any]] = []

    for recovered_table, loaded_table in zip(recovered_tables, loaded_tables, strict=False):
        row_start = int(recovered_table["row_start"])
        row_end = int(recovered_table["row_end"])
        column_start = int(recovered_table["column_start"])
        column_end = int(recovered_table["column_end"])
        if not (row_start <= row <= row_end and column_start <= column <= column_end):
            continue

        table_column_index = column - column_start + 1
        columns = list(recovered_table.get("columns") or [])
        db_columns = list(loaded_table.get("db_columns") or [])
        ref = {
            "source_name": str(recovered_table["name"]),
            "cell_role": "header" if row == int(recovered_table["header_row"]) else "data",
            "table_row_index": 0 if row == int(recovered_table["header_row"]) else row - int(recovered_table["header_row"]),
            "table_column_index": table_column_index,
            "column_name": str(columns[table_column_index - 1]) if table_column_index <= len(columns) else "",
            "db_column_name": str(db_columns[table_column_index - 1]) if table_column_index <= len(db_columns) else "",
        }
        refs.append(ref)

    return refs


def formulas_with_table_refs(
    formulas: list[dict[str, Any]],
    *,
    recovered_tables: list[dict[str, Any]],
    loaded_tables: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Attach table-relative overlays while preserving absolute formula pins."""
    annotated_formulas = []
    for formula in formulas:
        annotated_formula = deepcopy(formula)
        annotated_formula["table_refs"] = formula_table_refs(
            formula,
            recovered_tables=recovered_tables,
            loaded_tables=loaded_tables,
        )
        annotated_formulas.append(annotated_formula)
    return annotated_formulas
