"""Tabular file loading and bounded inspection helpers."""

from __future__ import annotations

from collections import Counter, deque
import csv
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook

MAX_SAMPLE_ROWS = 8
MAX_METADATA_ROWS = 20
MAX_SAMPLE_CHARS = 10000
MAX_ENCODING_SAMPLE_BYTES = 65536
MAX_FULL_PROFILE_BYTES = 50 * 1024 * 1024
MAX_FULL_EXTRACT_BYTES = 100 * 1024 * 1024


def tabular_dimensions(rows: list[list[str]]) -> tuple[int, int]:
    """Return the row and column counts for loaded rows."""
    return len(rows), max((len(row) for row in rows), default=0)


def tabular_summary(rows: list[list[str]], format_info: dict[str, Any]) -> dict[str, Any]:
    """Build the common file summary shared by all tabular tools."""
    row_count, column_count = tabular_dimensions(rows)
    return tabular_summary_from_counts(
        row_count=row_count,
        column_count=column_count,
        format_info=format_info,
    )


def tabular_summary_from_counts(
    *,
    row_count: int | None,
    column_count: int | None,
    format_info: dict[str, Any],
) -> dict[str, Any]:
    """Build a tabular summary from explicit row and column counts."""
    format_name = cast(str, format_info["format"])
    return {
        "format": format_name,
        "row_count": row_count,
        "column_count": column_count,
        **{key: value for key, value in format_info.items() if key != "format"},
    }


def preview_rows(rows: list[list[str]], sample_rows: int | None) -> list[list[str]]:
    """Return either all rows or a bounded preview."""
    return rows if sample_rows is None else rows[:sample_rows]


def preview_dimensions(rows: list[list[str]]) -> tuple[int, int]:
    """Return the row and column counts for a preview window."""
    return len(rows), max((len(row) for row in rows), default=0)


def count_non_empty(row: list[str]) -> int:
    """Count non-empty cells in a row."""
    return sum(bool(cell.strip()) for cell in row)


def is_blank(row: list[str]) -> bool:
    """Return whether a row has no non-empty cells."""
    return count_non_empty(row) == 0


def _normalize_cell(value: object) -> str:
    """Convert a raw cell value into the normalized string form."""
    if value is None:
        return ""
    return str(value)


def _read_file_sample(path: Path, *, sample_bytes: int) -> bytes:
    """Read at most ``sample_bytes`` from the start of a file."""
    with path.open("rb") as handle:
        return handle.read(sample_bytes)


def _csv_encoding(path: Path) -> str:
    """Detect a practical CSV text encoding from a small byte sample."""
    sample = _read_file_sample(path, sample_bytes=MAX_ENCODING_SAMPLE_BYTES)
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _csv_dialect(path: Path, *, encoding: str) -> csv.Dialect:
    """Detect a CSV dialect from a decoded sample."""
    with path.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(MAX_SAMPLE_CHARS)
    try:
        return csv.Sniffer().sniff(sample or ",")
    except csv.Error:
        return csv.get_dialect("excel")


def _csv_stream_info(path: Path) -> tuple[str, csv.Dialect, dict[str, Any]]:
    """Return the encoding, dialect, and common CSV format metadata."""
    encoding = _csv_encoding(path)
    dialect = _csv_dialect(path, encoding=encoding)
    return (
        encoding,
        dialect,
        {
            "format": "csv",
            "encoding": encoding,
            "delimiter": dialect.delimiter,
            "quotechar": dialect.quotechar,
            "sheet_names": [],
        },
    )


def _median_from_frequencies(frequencies: Counter[int], total_count: int) -> float:
    """Compute the median from a frequency map of integer values."""
    if total_count <= 0:
        return 0

    midpoint_low = (total_count - 1) // 2
    midpoint_high = total_count // 2
    running_total = 0
    low_value: int | None = None
    high_value = 0

    for value in sorted(frequencies):
        running_total += frequencies[value]
        if running_total > midpoint_low and low_value is None:
            low_value = value
        if running_total > midpoint_high:
            high_value = value
            break

    return float(((low_value or 0) + high_value) / 2)


def stream_csv_window(
    path: Path,
    *,
    start_row: int,
    limit: int,
    start_col: int,
    end_col: int | None,
) -> tuple[list[list[str]], dict[str, Any]]:
    """Stream a CSV once to collect a bounded grid window and file summary."""
    encoding, dialect, format_info = _csv_stream_info(path)
    selected_rows: list[list[str]] = []
    end_row = start_row + limit - 1

    with path.open("r", encoding=encoding, newline="") as handle:
        for row_index, row in enumerate(csv.reader(handle, dialect), start=1):
            if row_index < start_row:
                continue
            if row_index > end_row:
                break
            safe_end_col = len(row) if end_col is None else max(start_col, end_col)
            selected_rows.append(row[start_col - 1 : safe_end_col])

    return selected_rows, {
        **format_info,
        "read_mode": "bounded_stream",
    }


def stream_csv_profile(path: Path, *, max_sample_rows: int) -> dict[str, Any]:
    """Build a bounded CSV profile without materializing the full file."""
    encoding, dialect, format_info = _csv_stream_info(path)
    top_rows: list[list[str]] = []
    bottom_rows: deque[list[str]] = deque(maxlen=max_sample_rows)
    non_empty_frequencies: Counter[int] = Counter()
    row_count = 0
    column_count = 0
    non_empty_row_count = 0
    max_non_empty_cells = 0

    with path.open("r", encoding=encoding, newline="") as handle:
        for row in csv.reader(handle, dialect):
            row_count += 1
            column_count = max(column_count, len(row))
            if len(top_rows) < max_sample_rows:
                top_rows.append(list(row))
            bottom_rows.append(list(row))

            non_empty_cells = count_non_empty(row)
            if not non_empty_cells:
                continue
            non_empty_row_count += 1
            max_non_empty_cells = max(max_non_empty_cells, non_empty_cells)
            non_empty_frequencies[non_empty_cells] += 1

    sample_rows = top_rows
    tail_rows = list(bottom_rows) if row_count > max_sample_rows else []

    return {
        **format_info,
        "row_count": row_count,
        "column_count": column_count,
        "non_empty_row_count": non_empty_row_count,
        "blank_row_count": row_count - non_empty_row_count,
        "max_non_empty_cells_in_row": max_non_empty_cells,
        "median_non_empty_cells_per_non_blank_row": _median_from_frequencies(non_empty_frequencies, non_empty_row_count),
        "sample_rows": sample_rows,
        "header_candidates": [],
        "regions": [],
        "sheet_names": [],
        "profile_mode": "streaming",
        "top_rows": sample_rows,
        "bottom_rows": tail_rows,
    }


def _load_csv_rows(path: Path) -> tuple[list[list[str]], dict[str, str]]:
    """Load CSV rows and detected dialect metadata."""
    encoding, dialect, format_info = _csv_stream_info(path)
    with path.open("r", encoding=encoding, newline="") as handle:
        rows = [list(row) for row in csv.reader(handle, dialect)]
    return rows, {key: value for key, value in format_info.items() if key != "format"}


def _load_xlsx_rows(path: Path, *, sheet: str | None = None) -> tuple[list[list[str]], dict[str, Any]]:
    """Load worksheet rows and propagate merged-cell values."""
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        sheet_name = sheet or sheet_names[0]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Unknown worksheet '{sheet_name}' in {path.name}")
        worksheet = workbook[sheet_name]
        rows = [[_normalize_cell(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
        for merged_range in worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            value = rows[min_row - 1][min_col - 1]
            for row_index in range(min_row - 1, max_row):
                for column_index in range(min_col - 1, max_col):
                    rows[row_index][column_index] = value
        return rows, {"sheet_name": worksheet.title, "sheet_names": sheet_names}
    finally:
        workbook.close()


def load_rows(path: Path, *, sheet: str | None = None) -> tuple[list[list[str]], dict[str, Any]]:
    """Load tabular rows from a supported file type."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows, format_info = _load_csv_rows(path)
        return rows, {"format": "csv", **format_info}
    if suffix == ".xlsx":
        rows, format_info = _load_xlsx_rows(path, sheet=sheet)
        return rows, {"format": "xlsx", **format_info}
    raise ValueError(f"Unsupported tabular file type: {path.suffix}")
